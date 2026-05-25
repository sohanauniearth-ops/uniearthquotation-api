from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.drawing.image import Image as XLImg
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image as PILImage
import io, os, zipfile, tempfile, json
from lxml import etree

app = Flask(__name__)
CORS(app)

IMG_MAP = {}       # product_name -> png bytes
PRODUCT_LIST = []  # list of all product names from index (for matching)

def load_from_excel():
    global IMG_MAP, PRODUCT_LIST
    excel_path = os.path.join(os.path.dirname(__file__), 'Uniearth LPS Components Index (1).xlsx')
    if not os.path.exists(excel_path):
        print("ERROR: Excel file not found")
        return

    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            media = {}
            for f in z.namelist():
                if f.startswith('xl/media/'):
                    with z.open(f) as src:
                        media[os.path.basename(f)] = src.read()

            with z.open('xl/drawings/_rels/drawing1.xml.rels') as f:
                tree = etree.parse(f)
            rid_to_file = {r.get('Id'): r.get('Target').replace('../media/', '')
                          for r in tree.getroot()}

            with z.open('xl/drawings/drawing1.xml') as f:
                tree = etree.parse(f)
            root = tree.getroot()
            ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                  'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}

            row_to_img = {}
            for anchor in (root.findall('.//xdr:twoCellAnchor', ns) +
                          root.findall('.//xdr:oneCellAnchor', ns)):
                from_e = anchor.find('xdr:from', ns)
                if from_e is None: continue
                row = int(from_e.find('xdr:row', ns).text) + 1
                col = int(from_e.find('xdr:col', ns).text) + 1
                blip = anchor.find('.//a:blip', ns)
                if blip is None: continue
                rId = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                img_file = rid_to_file.get(rId, '')
                if col == 6 and img_file and row not in row_to_img:
                    row_to_img[row] = img_file

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Sheet1']

        for row_num in range(2, ws.max_row + 1):
            name = str(ws.cell(row=row_num, column=3).value or '').replace('\n', ' ').strip()
            remarks = str(ws.cell(row=row_num, column=4).value or '').replace('\n', ' ').strip()
            if not name:
                continue

            # Full name = "Product Name Remarks" or just "Product Name"
            full_name = f"{name} {remarks}".strip() if remarks else name
            PRODUCT_LIST.append(full_name)

            # Store image if this row has one
            img_file = row_to_img.get(row_num)
            if img_file and img_file in media:
                try:
                    pil = PILImage.open(io.BytesIO(media[img_file])).convert('RGB')
                    pil.thumbnail((80, 60), PILImage.LANCZOS)
                    buf = io.BytesIO()
                    pil.save(buf, 'PNG')
                    IMG_MAP[full_name] = buf.getvalue()
                    # Also store by name alone
                    IMG_MAP[name] = buf.getvalue()
                except Exception as e:
                    print(f"Image error row {row_num}: {e}")

        print(f"Loaded {len(PRODUCT_LIST)} products, {len(IMG_MAP)} with images")

    except Exception as e:
        print(f"Failed to load Excel: {e}")

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'products': len(PRODUCT_LIST), 'images': len(IMG_MAP)})

@app.route('/products', methods=['GET'])
def get_products():
    """Return all product names for the frontend dropdown"""
    return jsonify({'products': PRODUCT_LIST})

@app.route('/generate', methods=['POST', 'OPTIONS'])
def generate():
    if request.method == 'OPTIONS':
        return '', 204
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    try:
        excel_bytes = build_excel(
            data.get('items', []),
            data.get('client', 'Client'),
            data.get('project', 'Project'),
            data.get('qno', 'Q-001'),
            data.get('validity', 10),
            data.get('today', ''),
            data.get('is_sitc', False)
        )
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Uniearth_{data.get('qno','Q-001')}.xlsx"
        )
    except Exception as e:
        print(f"Generate error: {e}")
        return jsonify({'error': str(e)}), 500

def build_excel(items, client, project, qno, validity, today, is_sitc):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Quotation'

    thin = Side(style='thin', color='CCCCCC')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    alt_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')

    def c(r, col, v='', bold=False, size=9, align='left', fill=None, num_fmt=None):
        cell = ws.cell(row=r, column=col, value=v)
        cell.font = Font(name='Arial', bold=bold, size=size)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = bdr
        if fill: cell.fill = fill
        if num_fmt: cell.number_format = num_fmt
        return cell

    ws.merge_cells('A1:G1')
    ws['A1'] = 'Uniearth Earthing Solutions Pvt Ltd'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1B2A4A')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:G2')
    ws['A2'] = '3/35 West Punjabi Bagh, New Delhi - 110026  |  info@uniearth.co.in  |  www.uniearth.co.in  |  Ph: 011-45054847  |  Mob: +91-9811856550'
    ws['A2'].font = Font(name='Arial', size=8, italic=True, color='444444')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 13
    ws.row_dimensions[3].height = 5

    for row, l1, v1, l2, v2 in [
        (4, 'Company Name:', client, 'Ref. No.:', qno),
        (5, 'Project:', project, 'Date:', today),
        (6, 'Contact Person:', '', 'Ph. No.', ''),
        (7, 'E-Mail ID:', '', '', ''),
    ]:
        ws.cell(row=row, column=1, value=l1).font = Font(name='Arial', bold=True, size=9)
        ws.merge_cells(f'B{row}:D{row}')
        ws.cell(row=row, column=2, value=v1).font = Font(name='Arial', size=9)
        ws.cell(row=row, column=5, value=l2).font = Font(name='Arial', bold=True, size=9)
        ws.merge_cells(f'F{row}:G{row}')
        ws.cell(row=row, column=6, value=v2).font = Font(name='Arial', bold=True, size=9)
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[8].height = 5
    HDR = 9
    ws.row_dimensions[HDR].height = 30
    for i, h in enumerate(['S.No', 'Description of Item', 'Item\nImages', 'UOM', 'Qty', 'Rate\n(Supply)', 'Supply\nTotal'], 1):
        cell = ws.cell(row=HDR, column=i, value=h)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = bdr

    ROW_H = 55
    PT_TO_EMU = 12700
    CHAR_TO_EMU = 66675
    header_heights_pt = [26, 13, 5, 16, 16, 16, 16, 5, 30]
    sub = 0
    tmp_dir = tempfile.mkdtemp()

    for i, item in enumerate(items):
        r = HDR + 1 + i
        ws.row_dimensions[r].height = ROW_H
        fill = alt_fill if i % 2 == 0 else None

        # Description = matched product name from backend
        desc = item.get('product_name', item.get('desc', ''))
        c(r, 1, i + 1, align='center', fill=fill)
        c(r, 2, desc, fill=fill)
        c(r, 3, '', fill=fill)
        c(r, 4, item.get('unit', ''), align='center', fill=fill)
        c(r, 5, item.get('qty', 1), align='center', fill=fill)
        yp = round(item.get('your_price', 0))
        qty = item.get('qty', 1)
        c(r, 6, yp, align='right', fill=fill, num_fmt='#,##0')
        total = round(yp * qty)
        c(r, 7, total, align='right', fill=fill, num_fmt='#,##0', bold=True)
        sub += total

        # Image
        top_emu = (sum(h * PT_TO_EMU for h in header_heights_pt) +
                   i * ROW_H * PT_TO_EMU + int(0.15 * PT_TO_EMU))
        left_emu = int((6 + 52) * CHAR_TO_EMU) + int(0.2 * CHAR_TO_EMU)

        img_data = IMG_MAP.get(desc)
        if img_data:
            try:
                path = os.path.join(tmp_dir, f'img_{i}.png')
                with open(path, 'wb') as f:
                    f.write(img_data)
                xl_img = XLImg(path)
                xl_img.anchor = AbsoluteAnchor(
                    pos=XDRPoint2D(left_emu, top_emu),
                    ext=XDRPositiveSize2D(cm_to_EMU(2.2), cm_to_EMU(1.6))
                )
                ws.add_image(xl_img)
            except Exception as e:
                print(f"Image embed error row {i}: {e}")

    gst = round(sub * 0.18)
    last = HDR + 1 + len(items)
    ws.row_dimensions[last].height = 5

    for j, (label, val) in enumerate([('Sub Total', sub), ('IGST @ 18%', gst), ('Total Amount', sub + gst)], last + 1):
        ws.row_dimensions[j].height = 18
        ws.merge_cells(f'A{j}:E{j}')
        lc = ws.cell(row=j, column=1, value=label)
        lc.font = Font(name='Arial', bold=True, size=10)
        lc.alignment = Alignment(horizontal='right', vertical='center')
        lc.border = bdr
        vc = ws.cell(row=j, column=6, value=val)
        vc.font = Font(name='Arial', bold=True, size=10)
        vc.number_format = '#,##0'
        vc.alignment = Alignment(horizontal='right', vertical='center')
        vc.border = bdr
        ws.cell(row=j, column=7).border = bdr

    tr = last + 6
    ws.merge_cells(f'A{tr}:G{tr}')
    ws.cell(row=tr, column=1, value='TERMS & CONDITIONS').font = Font(name='Arial', bold=True, size=10)
    for j, t in enumerate([
        f'1. Validity Period: The above rates are valid for {validity} days from date of quotation.',
        '2. Taxation: All applicable taxes shall be charged extra as per actual at delivery time.',
        '3. Payment terms: 100% advance along with confirmed purchase order.',
        '4. Transportation cost: Freight will be charged extra as per actual.'
    ], tr + 1):
        ws.merge_cells(f'A{j}:G{j}')
        ws.cell(row=j, column=1, value=t).font = Font(name='Arial', size=9)
        ws.row_dimensions[j].height = 14

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

load_from_excel()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
